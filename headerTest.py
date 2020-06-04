import openpyxl


wb = openpyxl.Workbook()

sheet = wb.active

sheet.merge_cells('A1:D1')
sheet.merge_cells('F1:I1')
sheet.merge_cells('K1:N1')
sheet.merge_cells('P1:S1')
sheet.merge_cells('V1:Y1')
sheet.merge_cells('AA1:AD1')
sheet.merge_cells('AF1:AI1')
sheet.merge_cells('AK1:AN1')
sheet.merge_cells('AP1:AS1')
sheet.merge_cells('AV1:AY1')
sheet.merge_cells('BA1:BD1')
sheet.merge_cells('BF1:BI1')


sheet['A1'] = 'First Community'
sheet['A3'] = 'Lot No'
sheet['B3'] = 'Legal Lot'
sheet['C3'] = 'Block'
sheet['D3'] = 'Closing Date'

sheet['F1'] = 'Second Community'
sheet['F3'] = 'Lot No'
sheet['G3'] = 'Legal Lot'
sheet['H3'] = 'Block'
sheet['I3'] = 'Closing Date'

sheet['K1'] = 'Third Community'
sheet['K3'] = 'Lot No'
sheet['L3'] = 'Legal Lot'
sheet['M3'] = 'Block'
sheet['N3'] = 'Closing Date'


wb.save('Header Test.xlsx')